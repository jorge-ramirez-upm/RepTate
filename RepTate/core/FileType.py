# RepTate: Rheology of Entangled Polymers: Toolkit for the Analysis of Theory and Experiments
# --------------------------------------------------------------------------------------------------------
#
# Authors:
#     Jorge Ramirez, jorge.ramirez@upm.es
#     Victor Boudara, victor.boudara@gmail.com
#
# Useful links:
#     http://blogs.upm.es/compsoftmatter/software/reptate/
#     https://github.com/jorge-ramirez-upm/RepTate
#     http://reptate.readthedocs.io
#
# --------------------------------------------------------------------------------------------------------
#
# Copyright (2017-2026): Jorge Ramirez, Victor Boudara, Universidad Politécnica de Madrid, University of Leeds
#
# This file is part of RepTate.
#
# RepTate is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# RepTate is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with RepTate.  If not, see <http://www.gnu.org/licenses/>.
#
# --------------------------------------------------------------------------------------------------------
"""Module FileType

Module for the basic definition of file types.

"""
import os
from typing import Any, TypeAlias

import numpy as np

# import logging
from openpyxl import load_workbook
from RepTate.core.File import File, FileParameterSpec
from RepTate.core.typing import AxesArray
from RepTate.core.units import (
    convert_array_to_internal,
    make_column_specs,
    parse_column_label,
    parse_parameter_value,
)


FileParameterSpecs: TypeAlias = dict[str, FileParameterSpec]


class TXTColumnFile(object):
    """Basic class for text-column based data files

    Columns should be separated by espaces or tabs

    BASIC Structure of FILE::

        LINE    CONTENTS
        0       Param1=434;Param2=4355;
        1       # Header line and/or comments [OPTIONAL, ANY NUMBER OF HEADER LINES IS POSSIBLE]
        2       col1 col2 col3 [NAMES OF COLUMNS, OPTIONAL]
        3       4343 434.5 535e-434 [DATA, ONLY NUMBERS ALLOWED]

    The following examples can be declared with the line::

        ftype=TXTColumnFile("LVE files", "tts", "LVE files", ['w','G\'','G\'\''], ['Mw','T'], ['rad/s','Pa','Pa'])

    EXAMPLE 1: columns line, no header lines::

        C1=8.77210163229153;C2=114.03;Rho0=0.928;C3=0.61;T=-35;CTg=14.65;dx12=0;isof=true;Mw=634.5;chem=PI;PDI=1.03;
        w                           G'                          G''                        T                            g
        4.29882628773180E-0008      1.44001856995549E+0002      3.70207627600662E+0003     -3.30000000000000E-0003      0.00000000000000E+0000
        6.30767835406968E-0008      2.56947504513849E+0002      5.39032089470917E+0003      3.14760000000000E-0004      0.00000000000000E+0000
        9.25946098215800E-0008      4.87031807130633E+0002      7.86538338583378E+0003     -1.01000000000000E-0002      0.00000000000000E+0000

    EXAMPLE 2: Neither columns line nor header lines::

        C1=8.77210163229153;C2=114.03;Rho0=0.928;C3=0.61;T=-35;CTg=14.65;dx12=0;isof=true;Mw=23.4;chem=PI;PDI=1.03;
        2.42782390212358E-0003      2.11182193155015E+0001      1.72559181398615E+0003      1.25000000000000E-0003      0.00000000000000E+0000
        3.56351666244471E-0003      4.30476548641552E+0001      2.53490824331357E+0003      1.48400000000000E-0002      0.00000000000000E+0000

    EXAMPLE 3: 2 Header lines, no column line::

        T=160;chem=PE;
        # Header 1
        # Header 2
        4.23333e-05 1.05E+00 2.96E+01
        6.7e-05 2.02E+00 3.97E+01

    EXAMPLE 4: 2 Header lines + column line::

        T=160;chem=PE;
        # Header 1
        # Header 2
        w G' G''
        4.23333e-05 1.05E+00 2.96E+01
        6.7e-05 2.02E+00 3.97E+01

    """

    def __init__(
        self,
        name: str = "TXTColumn",
        extension: str = "txt",
        description: str = "Generic text file with columns",
        col_names: list[str] = [],
        basic_file_parameters: list[str] = [],
        col_units: list[str] = [],
        file_parameter_specs: list[FileParameterSpec] = [],
    ) -> None:
        """
        **Constructor**

        Keyword Arguments:
            - name {str} -- Name of file type
            - extension {str} -- File extension
            - description {str} -- Description of file contents
            - col_names {list of str}: list with names of columns to read
            - basic_file_parameters {list of str}: list with file parameters that should always be included in the header line
            - col_units {list of str}: Default units of columns
            - file_parameter_specs {list}: optional unit metadata for file parameters
        """
        self.name: str = name
        self.extension: str = extension
        self.description: str = description
        self.col_names_line: int = 0
        self.first_data_line: int = 0
        self.col_names: list[str] = col_names
        self.col_index: list[int] = list(range(len(self.col_names)))
        self.basic_file_parameters: list[str] = (
            basic_file_parameters  # Those that will show by default in the dataset
        )
        self.col_units: list[str] = col_units
        self.file_parameter_specs: FileParameterSpecs = {
            spec.name: spec for spec in file_parameter_specs
        }
        # self.logger = logging.getLogger('ReptateLogger')

    def is_number(self, s: str) -> bool:
        """Checks if the input string contains a number"""
        try:
            float(s)
            return True
        except ValueError:
            return False

    def get_parameters(self, line: str, file: File) -> None:
        """Get the file parameters"""
        items = line.split(";")
        file.file_parameters = {}
        for i in range(len(items)):
            par = items[i].split("=", 1)
            if len(par) > 1:
                name = par[0].strip()
                parsed_value, unit_symbol = parse_parameter_value(par[1].strip())
                if unit_symbol is not None and not isinstance(parsed_value, str):
                    file.set_file_parameter(
                        name,
                        parsed_value,
                        from_display=False,
                        source_unit=unit_symbol,
                    )
                elif isinstance(parsed_value, float):
                    file.set_file_parameter(name, parsed_value)
                else:
                    file.set_file_parameter(name, parsed_value)

    def find_col_names_and_first_data_lines(
        self, lines: list[str], file: File
    ) -> tuple[int, int]:
        """Find column names and first row with data"""
        colnameline = 0
        firstdata = 0
        for i in range(1, len(lines)):
            if all(x in lines[i] for x in self.col_names):
                # Column names line found
                colnameline = i
            elif all(self.is_number(x) for x in lines[i].split()):
                # Data lines have been found
                firstdata = i
                break
            else:
                # Otherwise, this must be a header line
                file.header_lines.append(lines[i])
        return colnameline, firstdata

    def parse_column_header(self, line: str) -> tuple[list[int], list[str], list[str]]:
        """Map expected columns to data indexes and units declared in a header.

        If no unit is declared in the file header, fall back to the application
        file type units. Those units are RepTate's current implicit input units.
        """
        tokens = line.split()
        col_index: list[int] = []
        col_labels: list[str] = []
        col_units: list[str] = []
        data_col_index = 0
        i = 0
        while i < len(tokens):
            token = tokens[i]
            unit_token = ""
            if i + 1 < len(tokens) and (
                tokens[i + 1].startswith("[") or tokens[i + 1].startswith("(")
            ):
                unit_token = tokens[i + 1]
                i += 1
            if token in self.col_names:
                default_unit = (
                    self.col_units[self.col_names.index(token)]
                    if self.col_names.index(token) < len(self.col_units)
                    else "-"
                )
                label, unit, _ = parse_column_label(
                    "%s %s" % (token, unit_token) if unit_token else token
                )
                col_index.append(data_col_index)
                col_labels.append(label)
                col_units.append(unit or default_unit)
            data_col_index += 1
            i += 1
        return col_index, col_labels, col_units

    def read_file(
        self, filename: str, parent_dataset: Any, axarr: AxesArray | None
    ) -> File | None:
        """Gets all the data from the file"""
        if not os.path.isfile(filename):
            print('File "%s" does not exists' % filename)  # pyright: ignore[reportUnboundVariable]
            return
        file = File(filename, self, parent_dataset, axarr)
        f = open(filename, "r", encoding="latin-1")
        lines = f.readlines()

        self.get_parameters(lines[0], file)
        (
            self.col_names_line,
            self.first_data_line,
        ) = self.find_col_names_and_first_data_lines(lines, file)

        self.col_index = []
        if self.col_names_line > 0:
            col_labels: list[str] = self.col_names[:]
            self.col_index, col_labels, col_units = self.parse_column_header(
                lines[self.col_names_line]
            )
        else:
            self.col_index = list(range(len(self.col_names)))
            col_labels = self.col_names[:]
            col_units = self.col_units[:]
        expected_units = [
            self.col_units[self.col_names.index(label)]
            if label in self.col_names and self.col_names.index(label) < len(self.col_units)
            else unit
            for label, unit in zip(col_labels, col_units)
        ]

        file.data_table.num_columns = len(self.col_index)
        file.data_table.column_names = col_labels
        file.data_table.column_units = col_units
        file.data_table.column_specs = make_column_specs(col_labels, col_units, expected_units)
        rawdata: list[float] = []
        for i in range(self.first_data_line, len(lines)):
            items = lines[i].split()
            if len(items) > 0:
                for j in self.col_index:
                    try:
                        rawdata.append(float(items[j]))
                    except (IndexError, ValueError):
                        rawdata.append(float("nan"))
        file.data_table.num_rows = int(len(rawdata) / file.data_table.num_columns)
        file.data_table.data = np.reshape(
            rawdata, newshape=(file.data_table.num_rows, file.data_table.num_columns)
        )
        file.data_table.data = file.data_table.data[
            file.data_table.data[:, 0].argsort()
        ]
        # TXTColumnFile units are application-declared implicit input/display
        # units. Known units are converted to RepTate's internal canonical units;
        # unknown legacy unit strings keep the existing numerical convention.
        for j, spec in enumerate(file.data_table.column_specs):
            file.data_table.data[:, j] = convert_array_to_internal(
                file.data_table.data[:, j], spec.display_unit, spec.internal_unit
            )

        return file


class ExcelFile(object):
    """Parse and read contents from Excel file"""

    def __init__(
        self,
        name: str = "Excel File",
        extension: str = "xlsx",
        description: str = "Generic Excel file",
        col_names: list[str] = [],
        basic_file_parameters: list[str] = [],
        col_units: list[str] = [],
        file_parameter_specs: list[FileParameterSpec] = [],
    ) -> None:

        self.name: str = name
        self.extension: str = extension
        self.description: str = description
        self.col_names: list[str] = col_names
        self.col_index: list[int] = list(range(len(self.col_names)))
        self.basic_file_parameters: list[str] = (
            basic_file_parameters  # Those that will show by default in the dataset
        )
        self.col_units: list[str] = col_units
        self.file_parameter_specs: FileParameterSpecs = {
            spec.name: spec for spec in file_parameter_specs
        }
        # self.logger = logging.getLogger('ReptateLogger')

    def read_file(
        self, filename: str, parent_dataset: Any, axarr: AxesArray | None
    ) -> File | None:
        """Read Excel File"""
        if not os.path.isfile(filename):
            print('File "%s" does not exists' % filename)  # pyright: ignore[reportUndefinedVariable]
            return
        file = File(filename, self, parent_dataset, axarr)
        wb = load_workbook(filename)
        for i, k in enumerate(wb.sheetnames):
            print("%d: %s" % (i, k))
        opt = int(
            input(
                "Select the Sheet that contains the data (number between 0 and %d) > "
                % (len(wb.sheetnames) - 1)
            )
        )
        if opt < 0 or opt >= len(wb.sheetnames):
            print("Invalid option!")
        ws = wb[wb.sheetnames[opt]]
        cexcelnames = ["A", "B", "C", "D", "E", "F"]
        for i in range(ws.max_column):
            print("%10s" % cexcelnames[i], end=" ")
        print("")
        for i, row in enumerate(ws.rows):
            for j, cell in enumerate(row):
                a = cell.value
                if type(a) is float:
                    print("%10.5g" % a, end=" ")
                elif type(a) is str:
                    print("%10s" % a, end=" ")
                elif type(a) is int:
                    print("%10d" % a, end=" ")
                if j > 10:
                    break
            print("")
            if i > 4:
                break
        file.data_table.num_rows = ws.max_row - 2
        file.data_table.num_columns = len(self.col_names)
        file.data_table.data = np.zeros(
            (file.data_table.num_rows, file.data_table.num_columns)
        )
        for j, n in enumerate(self.col_names):
            opt = ""
            while opt not in cexcelnames:
                opt = input("Column that contains the data for %s > " % n)
            for i in range(3, ws.max_row + 1):
                cell_name = "{}{}".format(opt, i)
                file.data_table.data[i - 3, j] = ws[cell_name].value
        return file
