# RepTate: Rheology of Entangled Polymers: Toolkit for the Analysis of Theory and Experiments
# --------------------------------------------------------------------------------------------------------
#
# Authors:
#     Jorge Ramirez, jorge.ramirez@upm.es
#     Victor Boudara, victor.boudara@gmail.com
#
# Useful links:
#     http://blogs.upm.es/compsoftmatter/software/reptate/
#     https://github.com/jorge-ramirez-upm/RepTate
#     http://reptate.readthedocs.io
#
# --------------------------------------------------------------------------------------------------------
#
# Copyright (2017-2020): Jorge Ramirez, Victor Boudara, Universidad Politécnica de Madrid, University of Leeds
#
# This file is part of RepTate.
#
# RepTate is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# RepTate is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with RepTate.  If not, see <http://www.gnu.org/licenses/>.
#
# --------------------------------------------------------------------------------------------------------
"""Module for importing data form Excel spreadsheets

""" 
import sys
import os
import numpy as np
from PyQt5 import QtGui
from PyQt5.uic import loadUiType
from PyQt5.QtCore import Qt, QItemSelectionModel
from PyQt5.QtWidgets import QApplication, QSizePolicy, QFileDialog, QTableWidgetItem, QTableWidget, QAbstractItemView, QMessageBox
from openpyxl import load_workbook
import xlrd

PATH = os.path.dirname(os.path.abspath(__file__))
Ui_ImportExcelMainWindow, QMainWindowImportExcel = loadUiType(os.path.join(PATH,'import_excel_dialog.ui'))

class ImportExcelWindow(QMainWindowImportExcel, Ui_ImportExcelMainWindow):
    list_AZ = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',
     'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ',
      'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA',
       'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ',
        'BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BY', 'BZ']
    MAX_ROW = 100
    MAX_COL = len(list_AZ)
    def __init__(self, parent=None, headers=["w", "G'", "G''"], file_param=["Mw", "T"]):
        super().__init__()
        self.setupUi(self)
        # self.show()
        self.filepath = ""
        self.dir_start = "~"
        self.is_xlsx = True
        self.wb = None
        self.sheet = None
        self.max_row = 0
        self.nskip = 0
        self.max_col = 0
        self.select_file_tb.clicked.connect(self.handle_get_file)
        self.skip_sb.valueChanged.connect(self.handle_nskip_changed)
        self.qtabs.currentChanged.connect(self.handle_tab_changed)
        self.col1_cb.activated.connect(self.handle_col1_cb_activated)
        self.col2_cb.activated.connect(self.handle_col2_cb_activated)
        self.col3_cb.activated.connect(self.handle_col3_cb_activated)
        
        self.headers = headers
        self.ncol = len(self.headers)
        self.file_param = file_param
        self.populate_file_param(file_param)
        self.update_cols_cb()
    
    def handle_col1_cb_activated(self):
        if self.wb == None:
            return
        sheet = self.qtabs.tabText(self.qtabs.currentIndex())
        table, selected_idx = self.qtables[sheet]
        selected_idx[0] = self.col1_cb.currentIndex()
        self.qtables[sheet] = [table, selected_idx]
        self.update_data_preview_table()
    
    def handle_col2_cb_activated(self):
        if self.wb == None:
            return
        sheet = self.qtabs.tabText(self.qtabs.currentIndex())
        table, selected_idx = self.qtables[sheet]
        selected_idx[1] = self.col2_cb.currentIndex()
        self.qtables[sheet] = [table, selected_idx]
        self.update_data_preview_table()
    
    def handle_col3_cb_activated(self):
        if self.wb == None:
            return
        sheet = self.qtabs.tabText(self.qtabs.currentIndex())
        table, selected_idx = self.qtables[sheet]
        selected_idx[2] = self.col3_cb.currentIndex()
        self.qtables[sheet] = [table, selected_idx]
        self.update_data_preview_table()

    def update_cols_cb(self):
        self.col1_cb.clear()
        self.col2_cb.clear()
        self.col1.setText("Select Column <b>%s</b>" % self.headers[0])
        self.col2.setText("Select Column <b>%s</b>" % self.headers[1])
        self.col1_cb.addItems(self.list_AZ[:self.max_col])
        self.col2_cb.addItems(self.list_AZ[:self.max_col])
        self.col2_cb.setCurrentIndex(1)
        if self.ncol > 2:
            self.col3_cb.clear()
            self.col3.setText("Select Column <b>%s</b>" % self.headers[2])
            self.col3_cb.addItems(self.list_AZ[:self.max_col])
            self.col3_cb.setCurrentIndex(2)
        else:
            self.col3.hide()
            self.col3_cb.hide()
    
    def handle_tab_changed(self, idx):
        table, selected_idx = self.qtables[self.qtabs.tabText(idx)]
        ncols = table.columnCount()
        self.col1_cb.clear()
        self.col2_cb.clear()
        self.col1_cb.addItems(self.list_AZ[:ncols])
        self.col2_cb.addItems(self.list_AZ[:ncols])
        self.col1_cb.setCurrentIndex(selected_idx[0])
        self.col2_cb.setCurrentIndex(selected_idx[1])
        if self.ncol > 2:
            self.col3_cb.clear()
            self.col3_cb.addItems(self.list_AZ[:ncols])
            self.col3_cb.setCurrentIndex(selected_idx[2])
        self.update_data_preview_table()

    def handle_nskip_changed(self):
        if self.wb == None:
            return
        self.nskip = self.skip_sb.value()
        self.update_data_preview_table()
    
    def col2num(self, col):
        num = 0
        for c in col:
            num = num * 26 + (ord(c) - ord('A')) + 1
        return num

    def update_data_preview_table(self):
        idx = self.qtabs.currentIndex()
        sname = self.qtabs.tabText(idx)
        col1 = self.col2num(self.col1_cb.currentText()) - 1
        col2 = self.col2num(self.col2_cb.currentText()) - 1 
        if self.ncol > 2:
            col3 = self.col2num(self.col3_cb.currentText()) - 1
        table, _ = self.qtables[sname]
        nrows = table.rowCount()
        ncols = table.columnCount()
        header_labels = [self.list_AZ[i] for i in range(ncols)]
        header_labels[col1] = self.headers[0]
        header_labels[col2] = self.headers[1]
        indexes = [table.model().index(k, col1) for k in range(self.nskip, nrows)]
        indexes += [table.model().index(k, col2) for k in range(self.nskip, nrows)]
        if self.ncol > 2:
            indexes += [table.model().index(k, col3) for k in range(self.nskip, nrows)]
            header_labels[col3] = self.headers[2]
        table.setHorizontalHeaderLabels(header_labels)
        flag = QItemSelectionModel.Select
        table.selectionModel().clearSelection()
        [table.selectionModel().select(i, flag) for i in indexes]
        table.setFocus()

    def get_data(self):
        x = []
        y = []
        z = []
        if self.wb == None:
            return (x, y, z, True)
        flag_nan = False
        col1 = self.col2num(self.col1_cb.currentText()) - 1
        col2 = self.col2num(self.col2_cb.currentText()) - 1 
        if self.ncol > 2:
            col3 = self.col2num(self.col3_cb.currentText()) - 1
        sname = self.qtabs.tabText(self.qtabs.currentIndex())
        if self.is_xlsx:
            sheet = self.wb[sname]
            max_row = sheet.max_row
            offset = 1
        else:
            sheet = self.wb.sheet_by_name(sname)
            max_row = sheet.nrows
            offset = 0
        for k in range(self.nskip, max_row):
            # x values
            if self.is_xlsx:
                cellx = sheet.cell(row=k + 1, column=col1 + 1)
            else:
                cellx = sheet.cell(k, col1)
            if hasattr(cellx, "value"):
                valx = cellx.value
            else:
                valx = ""
            try:
                x.append(float(valx))
            except (ValueError, TypeError):
                x.append(np.nan)
                flag_nan = True

            # y values
            if self.is_xlsx:
                celly = sheet.cell(row=k + 1, column=col2 + 1)
            else:
                celly = sheet.cell(k, col2)
            if hasattr(celly, "value"):
                valy = celly.value
            else:
                valy = ""
            try:
                y.append(float(valy))
            except (ValueError, TypeError):
                y.append(np.nan)
                flag_nan = True
            
            if len(self.headers) > 2:
                # y values
                if self.is_xlsx:
                    cellz = sheet.cell(row=k + 1, column=col3 + 1)
                else:
                    cellz = sheet.cell(k, col3)
                if hasattr(cellz, "value"):
                    valz = cellz.value
                else:
                    valz = ""
                try:
                    z.append(float(valz))
                except (ValueError, TypeError):
                    z.append(np.nan)
                    flag_nan = True
        res_dic = {"file": self.selected_file_label.text(), "sheet": sname, "x":x, "y":y, "z":z, "flag_nan": flag_nan, "col1": self.col1_cb.currentText(),"col2": self.col2_cb.currentText()}
        if len(self.headers) > 2:
            res_dic["col3"] = self.col3_cb.currentText()
        return res_dic

    def populate_file_param(self, params):
        self.file_param_txt.clear()
        txt = ""
        for p in params:
            txt += "%s=0;" % p
        self.file_param_txt.setText(txt)

    def handle_get_file(self):
        # file browser window  
        options = QFileDialog.Options()
        dilogue_name = "Select Excel Data File"
        ext_filter = "Excel file (*.xls *xlsx)"
        selected_file, _ = QFileDialog.getOpenFileName(
            self, dilogue_name, self.dir_start, ext_filter, options=options)
        self.handle_read_new_file(selected_file)

    def handle_read_new_file(self, path):
        if not os.path.isfile(path):
            return
        self.dir_start = os.path.dirname(path)
        self.qtabs.blockSignals(True)
        self.clear_tabs()
        fname = os.path.basename(path)
        self.is_xlsx = os.path.splitext(path)[-1] == ".xlsx"
        self.selected_file_label.setText(fname)
        self.filepath = path
        try:
            if self.is_xlsx:
                self.wb = load_workbook(filename=self.filepath, data_only=True)
                self.sheet_names = self.wb.sheetnames
            else:
                self.wb = xlrd.open_workbook(self.filepath)
                self.sheet_names = self.wb.sheet_names()
        except:
            # password protected?
            QMessageBox.warning(self, 'Open Excel File', 'Error: Could not read the Excel file.')
            return
        self.qtables = {}
        
        for sname in self.sheet_names:
            if self.is_xlsx:
                sheet = self.wb[sname]
                max_row = sheet.max_row
                max_col = sheet.max_column
            else:
                sheet = self.wb.sheet_by_name(sname)
                max_row = sheet.nrows
                max_col = sheet.ncols
            max_row = min(max_row, self.MAX_ROW)
            max_col = min(max_col, self.MAX_COL)
            qsheet = QTableWidget(max_row, max_col, self)
            qsheet.setSelectionMode(QAbstractItemView.NoSelection)
            for i in range(max_row):
                for j in range(max_col):
                    if self.is_xlsx:
                        cell = sheet.cell(row=i + 1, column=j + 1)
                    else:
                        cell = sheet.cell(i, j)
                    if hasattr(cell, "value"):
                        val = cell.value
                    else:
                        val = ""
                    item = QTableWidgetItem("%s" % val)
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    qsheet.setItem(i, j, item)
            self.qtabs.addTab(qsheet, sname)
            selected_cols = [0, 1, 2]
            self.qtables[sname] = [qsheet, selected_cols]
        self.qtabs.blockSignals(False)
        self.qtabs.blockSignals(False)
        self.qtabs.setCurrentIndex(0)
        self.handle_tab_changed(0)

    def clear_tabs(self):
        for _ in range(self.qtabs.count()):
            w = self.qtabs.widget(0)
            self.qtabs.removeTab(0)
            del w
        self.qtables = {}
        self.skip_sb.blockSignals(True)
        self.skip_sb.setValue(0)
        self.skip_sb.blockSignals(False)
        self.nskip = 0

    def dragEnterEvent(self, e):
        if e.mimeData().hasFormat('text/uri-list'):
            e.accept()
        else:
            e.ignore()
    
    def dropEvent(self, e):
        path = e.mimeData().urls()[0].toLocalFile()
        if os.path.splitext(path)[-1] == ".xls" or os.path.splitext(path)[-1] == ".xlsx":
            self.handle_read_new_file(path)
        else:
            pass
            # print("not a readable file")

if __name__ == '__main__':
    app = QApplication(sys.argv)
    GUI = Window()
    sys.exit(app.exec_())
